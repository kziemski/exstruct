import pandas as pd
import xlwings as xw
from pathlib import Path
import math

# Types
from xlwings import Book
from typing import List, Dict, Optional, Tuple
from .models.maps import *
from .models import *

from .utils import (
    detect_border_clusters,
    load_border_maps_xlsx,
    shrink_to_content_openpyxl,
    warn_once,
)
from .utils.parse_formula import parse_series_formula


# ===============================
# ユーティリティ関数
# ===============================
def compute_line_angle_deg(w: float, h: float) -> float:
    # Excel座標系（yは下向き）に合わせて 0=右, 90=下, 180=左, 270=上
    return math.degrees(math.atan2(h, w)) % 360.0


def angle_to_compass(angle: float) -> str:
    dirs = ["E", "SE", "S", "SW", "W", "NW", "N", "NE"]
    idx = int(((angle + 22.5) % 360) // 45)
    return dirs[idx]


def coord_to_cell_by_edges(row_edges, col_edges, x: float, y: float) -> Optional[str]:
    def find_index(edges, pos):
        for i in range(1, len(edges)):
            if edges[i - 1] <= pos < edges[i]:
                return i
        return None

    r = find_index(row_edges, y)
    c = find_index(col_edges, x)
    if r is None or c is None:
        return None
    return f"{xw.utils.col_name(c)}{r}"


def has_arrow(style_val) -> bool:
    """ArrowheadStyle が非ゼロ/None以外なら矢印ありとみなす簡易判定"""
    try:
        v = int(style_val)
        return v != 0
    except Exception:
        return False


# ===============================
# Excel 処理関数
# ===============================
def extract_sheet_cells(file_path: Path) -> Dict[str, List[CellRow]]:
    """Excel の各シートを文字列 dtype で読み込み、空セル除去して CellRow に変換"""
    dfs = pd.read_excel(file_path, header=None, sheet_name=None, dtype=str)
    result: Dict[str, List[CellRow]] = {}
    for sheet_name, df in dfs.items():
        df = df.fillna("")
        rows: List[CellRow] = []
        # itertuples(index=False) は0始まり → Excel行番号は +1
        for excel_row, row in enumerate(df.itertuples(index=False, name=None), start=1):
            filtered = {str(j): v for j, v in enumerate(row) if str(v).strip() != ""}
            if not filtered:
                continue
            rows.append(CellRow(r=excel_row, c=filtered))
        result[sheet_name] = rows
    return result


def iter_shapes_recursive(shp):
    """Group 内の shapes も含めて再帰的に yield する"""
    # 自身を返す
    yield shp

    # Group の場合のみ内部を再帰で処理
    try:
        if shp.api.Type == 6:  # msoGroup
            items = shp.api.GroupItems
            for i in range(1, items.Count + 1):
                inner = items.Item(i)
                # inner は COM オブジェクトなので、xlwings Shape に変換
                # xlwings の Shape には API 直接渡せないので Workbook から取得し直す必要あり
                try:
                    name = inner.Name
                    xl_shape = shp.parent.shapes[
                        name
                    ]  # 親シートの shapes から名前で取り直す
                except:
                    xl_shape = None

                if xl_shape is not None:
                    for s in iter_shapes_recursive(xl_shape):
                        yield s
    except Exception:
        pass


def get_shapes_with_position(workbook: Book) -> Dict[str, List[Shape]]:
    shape_data: Dict[str, List[Shape]] = {}
    for sheet in workbook.sheets:
        shapes: List[Shape] = []
        for shp in sheet.shapes:
            # 図形タイプ
            try:
                type_num = shp.api.Type
                shape_type_str = MSO_SHAPE_TYPE_MAP.get(
                    type_num, f"Unknown({type_num})"
                )
                # 除外
                if shape_type_str in ["Chart", "Comment", "Picture", "FormControl"]:
                    continue
                autoshape_type_str = None
                if type_num == 1:  # AutoShape
                    astype_num = shp.api.AutoShapeType
                    autoshape_type_str = MSO_AUTO_SHAPE_TYPE_MAP.get(
                        astype_num, f"Unknown({astype_num})"
                    )
            except Exception:
                type_num = None
                shape_type_str = None
                autoshape_type_str = None
            # テキスト
            try:
                text = shp.text.strip() if shp.text else ""
            except Exception:
                text = ""

            # AutoShape-Mixedかつテキストのないものを出力しない
            if autoshape_type_str in ["Mixed"] and text == "":
                continue

            shape_obj = Shape(
                text=text,
                l=int(shp.left),
                t=int(shp.top),
                w=int(shp.width) if shape_type_str == "Group" else None,
                h=int(shp.height) if shape_type_str == "Group" else None,
                type=f"{shape_type_str}{f"-{autoshape_type_str}" if autoshape_type_str else ""}",
            )
            try:
                # 1) 直線／コネクタ
                if type_num in (9, 3):  # msoLine=9, msoConnector=3
                    angle = compute_line_angle_deg(float(shp.width), float(shp.height))
                    shape_obj.angle_deg = angle
                    shape_obj.direction = angle_to_compass(angle)  # type: ignore
                    # rotation は line/connector のときだけ記録。0は出力しない
                    try:
                        rot = float(shp.api.Rotation)
                        if abs(rot) > 1e-6:
                            shape_obj.rotation = rot
                    except Exception:
                        pass
                    # ArrowheadStyle 取得（トークンを抑えたいなら省略可）
                    begin_style = None
                    end_style = None
                    try:
                        begin_style = int(shp.api.Line.BeginArrowheadStyle)
                        end_style = int(shp.api.Line.EndArrowheadStyle)
                        shape_obj.begin_arrow_style = begin_style
                        shape_obj.end_arrow_style = end_style
                    except Exception:
                        pass
                # 2) AutoShape の矢印（例: RightArrowなど）
                elif type_num == 1 and (
                    autoshape_type_str and "Arrow" in autoshape_type_str
                ):
                    # rotation は出力しない（内々で推定に使用）
                    try:
                        rot = float(shp.api.Rotation)
                    except Exception:
                        rot = 0.0
                else:
                    # その他の図形は rotation 等を記録しない
                    pass
            except Exception:
                # 小さな失敗は握りつぶして形状だけ残す
                pass
            shapes.append(shape_obj)
        shape_data[sheet.name] = shapes
    return shape_data


def integrate_shapes_into_json(
    cell_data: Dict[str, List[CellRow]],
    shape_data: Dict[str, List[Shape]],
    workbook: Book,
) -> Dict[str, SheetData]:
    """セルデータと図形を統合、図形の中心座標から近いセルに紐付け"""
    result: Dict[str, SheetData] = {}
    for sheet_name, rows in cell_data.items():
        sheet_shapes = shape_data.get(sheet_name, [])
        sheet = workbook.sheets[sheet_name]

        #  シート全体の行高さから累積座標を作成し、図形のY座標→行インデックスを推定する
        used = sheet.used_range
        max_row = used.last_cell.row
        row_edges: List[float] = [0.0]
        for r in range(1, max_row + 1):
            h = sheet.range((r, 1)).height
            row_edges.append(row_edges[-1] + float(h))

        sheet_model = SheetData(
            rows=rows,
            shapes=sheet_shapes,
            charts=get_charts(sheet),
            tables=detect_tables(sheet),
        )

        result[sheet_name] = sheet_model
    return result


def get_charts(sheet: xw.Sheet) -> List[Chart]:
    charts: List[Chart] = []
    for ch in sheet.charts:
        series_list: List[ChartSeries] = []
        y_axis_title: str = ""
        y_axis_range: List[int] = []
        error: Optional[str] = None

        try:
            chart_com = sheet.api.ChartObjects(ch.name).Chart
            chart_type_num = chart_com.ChartType  # 型番号（整数値）
            chart_type_label = XL_CHART_TYPE_MAP.get(
                chart_type_num, f"unknown_{chart_type_num}"
            )

            for s in chart_com.SeriesCollection():
                parsed = parse_series_formula(getattr(s, "Formula", ""))
                name_range = parsed["name_range"] if parsed else None
                x_range = parsed["x_range"] if parsed else None
                y_range = parsed["y_range"] if parsed else None

                series_list.append(
                    ChartSeries(
                        name=s.Name,
                        name_range=name_range,
                        x_range=x_range,
                        y_range=y_range,
                    )
                )

            # Y軸タイトルと範囲
            try:
                y_axis = chart_com.Axes(2, 1)
                if y_axis.HasTitle:
                    y_axis_title = y_axis.AxisTitle.Text
                y_axis_range = [y_axis.MinimumScale, y_axis.MaximumScale]
            except Exception:
                y_axis_title = ""
                y_axis_range = []

            title = chart_com.ChartTitle.Text if chart_com.HasTitle else None
        except Exception:
            print("グラフデータの解析に失敗しました。構造化未対応のグラフです。")
            title = None
            error = "json構造化に失敗しました"

        charts.append(
            Chart(
                name=ch.name,
                chart_type=chart_type_label,
                title=title,
                y_axis_title=y_axis_title,
                y_axis_range=y_axis_range,  # type: ignore
                series=series_list,
                l=int(ch.left),
                t=int(ch.top),
                error=error,
            )
        )
    return charts


def shrink_to_content(
    sheet: xw.Sheet,
    top: int,
    left: int,
    bottom: int,
    right: int,
    require_inside_border: bool = False,
    min_nonempty_ratio: float = 0.0,  # 端の列/行を残すために必要な非空率（0.0なら無効）
) -> Tuple[int, int, int, int]:
    """
    罫線クラスタから得た矩形(top,left,bottom,right)を、
    セル内容（値）と内側罫線の有無で外側の空白列・行をトリミングする。
    """
    rng = sheet.range((top, left), (bottom, right))
    vals = rng.value
    if vals is None:
        vals = []
    # 2D 正規化
    if not isinstance(vals, list):
        vals = [[vals]]
    elif vals and not isinstance(vals[0], list):
        vals = [vals]
    rows_n = len(vals)
    cols_n = len(vals[0]) if rows_n else 0

    def to_str(x):
        # xlwingsは数値/日付/None/文字列が混在するので安全に文字列化
        return "" if x is None else str(x)

    def is_empty_value(x):
        s = to_str(x).strip()  # whitespace を空扱い
        return s == ""

    def row_empty(i: int) -> bool:
        return cols_n == 0 or all(is_empty_value(vals[i][j]) for j in range(cols_n))

    def col_empty(j: int) -> bool:
        return rows_n == 0 or all(is_empty_value(vals[i][j]) for i in range(rows_n))

    def row_nonempty_ratio(i: int) -> float:
        if cols_n == 0:
            return 0.0
        cnt = sum(1 for j in range(cols_n) if not is_empty_value(vals[i][j]))
        return cnt / cols_n

    def col_nonempty_ratio(j: int) -> float:
        if rows_n == 0:
            return 0.0
        cnt = sum(1 for i in range(rows_n) if not is_empty_value(vals[i][j]))
        return cnt / rows_n

    # 内側罫線の定数
    XL_LINESTYLE_NONE = -4142
    XL_INSIDE_VERTICAL = 11
    XL_INSIDE_HORIZONTAL = 12

    def column_has_inside_border(col_idx: int) -> bool:
        # require_inside_border=False のときはチェック無効
        if not require_inside_border:
            return False
        try:
            for r in range(top, bottom + 1):
                ls = (
                    sheet.api.Cells(r, left + col_idx)
                    .Borders(XL_INSIDE_VERTICAL)
                    .LineStyle
                )
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    return True
        except Exception:
            pass
        return False

    def row_has_inside_border(row_idx: int) -> bool:
        if not require_inside_border:
            return False
        try:
            for c in range(left, right + 1):
                ls = (
                    sheet.api.Cells(top + row_idx, c)
                    .Borders(XL_INSIDE_HORIZONTAL)
                    .LineStyle
                )
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    return True
        except Exception:
            pass
        return False

    def should_trim_col(j: int) -> bool:
        # 空列 もしくは（内側罫線がない OR 非空率が閾値未満）
        if col_empty(j):
            return True
        if require_inside_border and not column_has_inside_border(j):
            return True
        if min_nonempty_ratio > 0.0 and col_nonempty_ratio(j) < min_nonempty_ratio:
            return True
        return False

    def should_trim_row(i: int) -> bool:
        if row_empty(i):
            return True
        if require_inside_border and not row_has_inside_border(i):
            return True
        if min_nonempty_ratio > 0.0 and row_nonempty_ratio(i) < min_nonempty_ratio:
            return True
        return False

    # 左
    while left <= right and cols_n > 0:
        if should_trim_col(0):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(0)
            cols_n = len(vals[0]) if rows_n else 0
            left += 1
        else:
            break
    # 上
    while top <= bottom and rows_n > 0:
        if should_trim_row(0):
            vals.pop(0)
            rows_n = len(vals)
            top += 1
        else:
            break
    # 右
    while left <= right and cols_n > 0:
        if should_trim_col(cols_n - 1):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(cols_n - 1)
            cols_n = len(vals[0]) if rows_n else 0
            right -= 1
        else:
            break
    # 下
    while top <= bottom and rows_n > 0:
        if should_trim_row(rows_n - 1):
            vals.pop(rows_n - 1)
            rows_n = len(vals)
            bottom -= 1
        else:
            break
    return top, left, bottom, right


def detect_tables_xlwings(sheet: xw.Sheet) -> List[str]:
    """
    罫線の連続領域から「表の塊」を検出するロジック（xlwingsによるCOM参照）
    1) まず ListObjects（正式な「テーブル」）を検出
    2) 次にセル罫線の4辺（＋必要なら内側）を見てクラスタ検出
    """
    tables: List[str] = []
    # 1) まずは Excel の正式なテーブル(ListObject)を拾う
    try:
        for lo in sheet.api.ListObjects:
            rng = lo.Range
            top_row = int(rng.Row)
            left_col = int(rng.Column)
            bottom_row = top_row + int(rng.Rows.Count) - 1
            right_col = left_col + int(rng.Columns.Count) - 1
            addr = rng.Address(RowAbsolute=False, ColumnAbsolute=False)
            tables.append(addr)
    except Exception:
        # ListObjects がない/取得失敗は無視して次へ
        pass
    # 2) フォールバック: 罫線クラスタ検出
    used = sheet.used_range
    max_row = used.last_cell.row
    max_col = used.last_cell.column
    # xlBordersIndex の定数（環境依存のため数値を直書き）
    # 参考: https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
    XL_DIAGONAL_DOWN = 5
    XL_DIAGONAL_UP = 6
    XL_EDGE_LEFT = 7
    XL_EDGE_TOP = 8
    XL_EDGE_BOTTOM = 9
    XL_EDGE_RIGHT = 10
    XL_INSIDE_VERTICAL = 11
    XL_INSIDE_HORIZONTAL = 12
    # 「罫線なし」判定に使う定数
    XL_LINESTYLE_NONE = -4142

    def cell_has_any_border(r: int, c: int) -> bool:
        """セルに外枠/内枠いずれかの可視罫線があるか粗く判定"""
        try:
            b = sheet.api.Cells(r, c).Borders
            # 4辺＋内側（必要に応じて内側は除外してもOK）
            for idx in (
                XL_EDGE_LEFT,
                XL_EDGE_TOP,
                XL_EDGE_RIGHT,
                XL_EDGE_BOTTOM,
                XL_INSIDE_VERTICAL,
                XL_INSIDE_HORIZONTAL,
            ):
                ls = b(idx).LineStyle
                # 罫線なし: xlLineStyleNone(-4142)。0 は環境によっては出ないことがある
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    # 追加の堅牢性: Weight も見る（0 は線なし扱い）
                    try:
                        if getattr(b(idx), "Weight", 0) == 0:
                            continue
                    except Exception:
                        pass
                    return True
            return False
        except Exception:
            return False

    # セルごとの True/False マップ
    grid = [[False] * (max_col + 1) for _ in range(max_row + 1)]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if cell_has_any_border(r, c):
                grid[r][c] = True
    visited = [[False] * (max_col + 1) for _ in range(max_row + 1)]

    def dfs(sr: int, sc: int, acc: List[Tuple[int, int]]):
        stack = [(sr, sc)]
        while stack:
            rr, cc = stack.pop()
            if not (1 <= rr <= max_row and 1 <= cc <= max_col):
                continue
            if visited[rr][cc] or not grid[rr][cc]:
                continue
            visited[rr][cc] = True
            acc.append((rr, cc))
            for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                stack.append((rr + dr, cc + dc))

    clusters: List[Tuple[int, int, int, int]] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if grid[r][c] and not visited[r][c]:
                cluster: List[Tuple[int, int]] = []
                dfs(r, c, cluster)
                if len(cluster) < 4:
                    # ノイズを除外（閾値は適宜調整）
                    continue
                rows = [rc[0] for rc in cluster]
                cols = [rc[1] for rc in cluster]
                top_row = min(rows)
                bottom_row = max(rows)
                left_col = min(cols)
                right_col = max(cols)
                clusters.append((top_row, left_col, bottom_row, right_col))

    # 重複や包含の矩形をまとめる（簡易マージ）
    def overlaps(a, b):
        return not (a[1] > b[3] or a[3] < b[1] or a[0] > b[2] or a[2] < b[0])

    merged_rects: List[Tuple[int, int, int, int]] = []
    for rect in sorted(clusters):
        merged = False
        for i, ex in enumerate(merged_rects):
            if overlaps(rect, ex):
                # 簡易的に外接矩形でマージ
                merged_rects[i] = (
                    min(rect[0], ex[0]),
                    min(rect[1], ex[1]),
                    max(rect[2], ex[2]),
                    max(rect[3], ex[3]),
                )
                merged = True
                break
        if not merged:
            merged_rects.append(rect)

    for top_row, left_col, bottom_row, right_col in merged_rects:
        # 内容でトリミング（内側罫線チェックは必要に応じて True に）
        top_row, left_col, bottom_row, right_col = shrink_to_content(
            sheet, top_row, left_col, bottom_row, right_col, require_inside_border=False
        )
        addr = f"{xw.utils.col_name(left_col)}{top_row}:{xw.utils.col_name(right_col)}{bottom_row}"
        tables.append(addr)
    return tables


def detect_tables_openpyxl(xlsx_path: Path, sheet_name: str) -> List[str]:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    tables: List[str] = []
    # 1) まずは正式テーブル（ListObject相当）を拾う
    try:
        # openpyxl 3.x: ws.tables は dict-like or list-like
        openpyxl_tables = []
        if hasattr(ws, "tables") and ws.tables:
            # ws.tables が dict の場合（openpyxlのバージョンにより）
            if isinstance(ws.tables, dict):
                openpyxl_tables = list(ws.tables.values())
            else:
                openpyxl_tables = list(ws.tables)
        elif hasattr(ws, "_tables") and ws._tables:  # type: ignore
            openpyxl_tables = list(ws._tables)  # type: ignore
        for t in openpyxl_tables:
            # t.ref は 'A1:D8' のような範囲文字列
            addr = t.ref
            tables.append(addr)
    except Exception:
        pass
    # 2) 罫線クラスタ検出
    has_border, top_edge, bottom_edge, left_edge, right_edge, max_row, max_col = (
        load_border_maps_xlsx(xlsx_path, sheet_name)
    )
    rects = detect_border_clusters(has_border, min_size=4)

    # 矩形マージ（既存の overlaps と同じロジック）
    def overlaps(a, b):
        return not (a[1] > b[3] or a[3] < b[1] or a[0] > b[2] or a[2] < b[0])

    merged_rects: List[Tuple[int, int, int, int]] = []
    for rect in sorted(rects):
        merged = False
        for i, ex in enumerate(merged_rects):
            if overlaps(rect, ex):
                merged_rects[i] = (
                    min(rect[0], ex[0]),
                    min(rect[1], ex[1]),
                    max(rect[2], ex[2]),
                    max(rect[3], ex[3]),
                )
                merged = True
                break
        if not merged:
            merged_rects.append(rect)

    # トリミングして最終テーブルを追加
    for top_row, left_col, bottom_row, right_col in merged_rects:
        top_row, left_col, bottom_row, right_col = shrink_to_content_openpyxl(
            ws,
            top_row,
            left_col,
            bottom_row,
            right_col,
            require_inside_border=False,  # 必要ならTrueに
            top_edge=top_edge,
            bottom_edge=bottom_edge,
            left_edge=left_edge,
            right_edge=right_edge,
            min_nonempty_ratio=0.0,
        )
        addr = f"{get_column_letter(left_col)}{top_row}:{get_column_letter(right_col)}{bottom_row}"
        tables.append(addr)
    wb.close()
    return tables


def detect_tables(sheet: xw.Sheet) -> List[str]:
    """
    .xlsx/.xlsm は openpyxl で高速検出。
    .xls や openpyxl が使えない場合は xlwings 版にフォールバック。
    """
    excel_path: Optional[Path] = None
    try:
        excel_path = Path(sheet.book.fullname)
    except Exception:
        excel_path = None

    # .xls（バイナリ）は openpyxl 非対応 → フォールバック＋警告
    if excel_path and excel_path.suffix.lower() == ".xls":
        warn_once(
            f"xls-fallback::{excel_path}",
            f"File '{excel_path.name}' is .xls (BIFF); openpyxl cannot read it. Falling back to COM-based detection (slower). Consider converting to .xlsx.",
        )
        return detect_tables_xlwings(sheet)

    # .xlsx / .xlsm なら openpyxl 優先
    if excel_path and excel_path.suffix.lower() in (".xlsx", ".xlsm"):
        # openpyxl 自体が未インストールのケース
        try:
            import openpyxl  # noqa: F401
        except Exception:
            warn_once(
                "openpyxl-missing",
                "openpyxl is not installed. Falling back to COM-based detection (slower).",
            )
            return detect_tables_xlwings(sheet)

        # openpyxl 解析自体が失敗したケース
        try:
            return detect_tables_openpyxl(excel_path, sheet.name)
        except Exception as e:
            warn_once(
                f"openpyxl-parse-fallback::{excel_path}::{sheet.name}",
                f"openpyxl failed to parse '{excel_path.name}' (sheet '{sheet.name}'): {e!r}. Falling back to COM-based detection (slower).",
            )
            return detect_tables_xlwings(sheet)

    # 3) パスが取れない等の不明な場合もフォールバック
    warn_once(
        "unknown-ext-fallback",
        "Workbook path or extension is unavailable; falling back to COM-based detection (slower).",
    )
    return detect_tables_xlwings(sheet)


def dict_without_empty_values(obj):
    """モデルや辞書から空値（None, '', [], {}}）を除外する再帰関数"""
    if isinstance(obj, dict):
        return {
            k: dict_without_empty_values(v)
            for k, v in obj.items()
            if v not in [None, "", [], {}]
        }
    elif isinstance(obj, list):
        return [
            dict_without_empty_values(v) for v in obj if v not in [None, "", [], {}]
        ]
    elif hasattr(obj, "model_dump"):
        # pydanticモデルの場合
        return dict_without_empty_values(obj.model_dump(exclude_none=True))
    else:
        return obj


# ===============================
# 出力関数
# ===============================
def save_as_json(model: WorkbookData, path: Path):
    # 空値除外してからjson化
    filtered_dict = dict_without_empty_values(model)
    import json

    path.write_text(
        json.dumps(filtered_dict, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ===============================
# 単一 Excel ファイル処理
# ===============================
def process_excel(
    file_path: Path, output_path: Path, out_fmt: str, image: bool, pdf: bool, dpi: int
):
    cell_data = extract_sheet_cells(file_path)
    wb = xw.Book(file_path)
    try:
        shape_data = get_shapes_with_position(wb)
        merged = integrate_shapes_into_json(cell_data, shape_data, wb)
    finally:
        wb.close()

    workbook_model = WorkbookData(book_name=file_path.name, sheets=merged)

    if out_fmt == "json":
        save_as_json(workbook_model, output_path)

    print(f"{file_path.name} → {output_path} 完了")
