from __future__ import annotations

from collections.abc import Callable, Sequence
from dataclasses import dataclass, field
import logging
import os
from pathlib import Path
from typing import Literal

import xlwings as xw

from ..models import CellRow, PrintArea, SheetData, WorkbookData
from .cells import (
    WorkbookColorsMap,
    detect_tables_openpyxl,
    extract_sheet_cells,
    extract_sheet_cells_with_links,
    extract_sheet_colors_map,
    extract_sheet_colors_map_com,
)

logger = logging.getLogger(__name__)

ExtractionMode = Literal["light", "standard", "verbose"]
CellData = dict[str, list[CellRow]]
PrintAreaData = dict[str, list[PrintArea]]


@dataclass(frozen=True)
class ExtractionInputs:
    """Immutable inputs for pipeline steps.

    Attributes:
        file_path: Path to the Excel workbook.
        mode: Extraction mode (light/standard/verbose).
        include_cell_links: Whether to include cell hyperlinks.
        include_print_areas: Whether to include print areas.
        include_auto_page_breaks: Whether to include auto page breaks.
        include_colors_map: Whether to include background colors map.
        include_default_background: Whether to include default background color.
        ignore_colors: Optional set of color keys to ignore.
    """

    file_path: Path
    mode: ExtractionMode
    include_cell_links: bool
    include_print_areas: bool
    include_auto_page_breaks: bool
    include_colors_map: bool
    include_default_background: bool
    ignore_colors: set[str] | None


@dataclass
class ExtractionArtifacts:
    """Mutable artifacts collected by pipeline steps.

    Attributes:
        cell_data: Extracted cell rows per sheet.
        print_area_data: Extracted print areas per sheet.
        auto_page_break_data: Extracted auto page-break areas per sheet.
        colors_map_data: Extracted colors map for workbook sheets.
    """

    cell_data: CellData = field(default_factory=dict)
    print_area_data: PrintAreaData = field(default_factory=dict)
    auto_page_break_data: PrintAreaData = field(default_factory=dict)
    colors_map_data: WorkbookColorsMap | None = None


ExtractionStep = Callable[[ExtractionInputs, ExtractionArtifacts], None]


def build_pre_com_pipeline(inputs: ExtractionInputs) -> list[ExtractionStep]:
    """Build pipeline steps that run before COM/Excel access.

    Args:
        inputs: Pipeline inputs describing extraction flags.

    Returns:
        Ordered list of extraction steps to run before COM.
    """
    steps: list[ExtractionStep] = [step_extract_cells]
    if inputs.include_print_areas:
        steps.append(step_extract_print_areas_openpyxl)
    if inputs.include_colors_map and (
        inputs.mode == "light" or os.getenv("SKIP_COM_TESTS")
    ):
        steps.append(step_extract_colors_map_openpyxl)
    return steps


def run_pipeline(
    steps: Sequence[ExtractionStep],
    inputs: ExtractionInputs,
    artifacts: ExtractionArtifacts,
) -> ExtractionArtifacts:
    """Run steps in order and return updated artifacts.

    Args:
        steps: Ordered extraction steps.
        inputs: Pipeline inputs.
        artifacts: Artifact container to update.

    Returns:
        Updated artifacts after running all steps.
    """
    for step in steps:
        step(inputs, artifacts)
    return artifacts


def step_extract_cells(
    inputs: ExtractionInputs, artifacts: ExtractionArtifacts
) -> None:
    """Extract cell rows, optionally including hyperlinks.

    Args:
        inputs: Pipeline inputs.
        artifacts: Artifact container to update.
    """
    artifacts.cell_data = (
        extract_sheet_cells_with_links(inputs.file_path)
        if inputs.include_cell_links
        else extract_sheet_cells(inputs.file_path)
    )


def step_extract_print_areas_openpyxl(
    inputs: ExtractionInputs, artifacts: ExtractionArtifacts
) -> None:
    """Extract print areas via openpyxl.

    Args:
        inputs: Pipeline inputs.
        artifacts: Artifact container to update.
    """
    artifacts.print_area_data = extract_print_areas_openpyxl(inputs.file_path)


def step_extract_colors_map_openpyxl(
    inputs: ExtractionInputs, artifacts: ExtractionArtifacts
) -> None:
    """Extract colors_map via openpyxl; logs and skips on failure.

    Args:
        inputs: Pipeline inputs.
        artifacts: Artifact container to update.
    """
    try:
        artifacts.colors_map_data = extract_sheet_colors_map(
            inputs.file_path,
            include_default_background=inputs.include_default_background,
            ignore_colors=inputs.ignore_colors,
        )
    except Exception as exc:
        logger.warning("Color map extraction failed; skipping colors_map. (%r)", exc)
        artifacts.colors_map_data = None


def extract_print_areas_openpyxl(file_path: Path) -> PrintAreaData:
    """Extract print areas per sheet using openpyxl defined names.

    Args:
        file_path: Path to the workbook.

    Returns:
        Mapping of sheet name to print area list.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        areas = _extract_print_areas_from_defined_names(wb)
        if not areas:
            areas = _extract_print_areas_from_sheet_props(wb)
        return areas
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _extract_print_areas_from_defined_names(workbook: object) -> PrintAreaData:
    """Extract print areas from defined names in an openpyxl workbook.

    Args:
        workbook: openpyxl workbook instance.

    Returns:
        Mapping of sheet name to print area list.
    """
    defined = getattr(workbook, "defined_names", None)
    if defined is None:
        return {}
    defined_area = defined.get("_xlnm.Print_Area")
    if not defined_area:
        return {}

    areas: PrintAreaData = {}
    sheetnames = set(getattr(workbook, "sheetnames", []))
    for sheet_name, range_str in defined_area.destinations:
        if sheet_name not in sheetnames:
            continue
        _append_print_areas(areas, sheet_name, str(range_str))
    return areas


def _extract_print_areas_from_sheet_props(workbook: object) -> PrintAreaData:
    """Extract print areas from sheet-level print area properties.

    Args:
        workbook: openpyxl workbook instance.

    Returns:
        Mapping of sheet name to print area list.
    """
    areas: PrintAreaData = {}
    worksheets = getattr(workbook, "worksheets", [])
    for ws in worksheets:
        pa = getattr(ws, "_print_area", None)
        if not pa:
            continue
        _append_print_areas(areas, str(getattr(ws, "title", "")), str(pa))
    return areas


def _append_print_areas(areas: PrintAreaData, sheet_name: str, range_str: str) -> None:
    """Append parsed print areas to the mapping.

    Args:
        areas: Mapping to update.
        sheet_name: Target sheet name.
        range_str: Raw range string, possibly comma-separated.
    """
    for part in str(range_str).split(","):
        parsed = _parse_print_area_range(part)
        if not parsed:
            continue
        r1, c1, r2, c2 = parsed
        areas.setdefault(sheet_name, []).append(PrintArea(r1=r1, c1=c1, r2=r2, c2=c2))


def _parse_print_area_range(range_str: str) -> tuple[int, int, int, int] | None:
    """Parse an Excel range string into zero-based coordinates.

    Args:
        range_str: Excel range string.

    Returns:
        Zero-based (r1, c1, r2, c2) tuple or None on failure.
    """
    from openpyxl.utils import range_boundaries

    cleaned = range_str.strip()
    if not cleaned:
        return None
    if "!" in cleaned:
        cleaned = cleaned.split("!", 1)[1]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    except Exception:
        return None
    return (min_row - 1, min_col - 1, max_row - 1, max_col - 1)


def extract_print_areas_com(workbook: xw.Book) -> PrintAreaData:
    """Extract print areas per sheet via xlwings/COM.

    Args:
        workbook: xlwings workbook.

    Returns:
        Mapping of sheet name to print area list.
    """
    from openpyxl.utils import range_boundaries

    def parse_range(range_str: str) -> tuple[int, int, int, int] | None:
        cleaned = range_str.strip()
        if not cleaned:
            return None
        if "!" in cleaned:
            cleaned = cleaned.split("!", 1)[1]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cleaned)
        except Exception:
            return None
        return (min_row - 1, min_col - 1, max_row - 1, max_col - 1)

    areas: PrintAreaData = {}
    for sheet in workbook.sheets:
        try:
            raw = sheet.api.PageSetup.PrintArea or ""
        except Exception:
            continue
        if not raw:
            continue
        for part in str(raw).split(","):
            parsed = parse_range(part)
            if not parsed:
                continue
            r1, c1, r2, c2 = parsed
            areas.setdefault(sheet.name, []).append(
                PrintArea(r1=r1, c1=c1, r2=r2, c2=c2)
            )
    return areas


def extract_colors_map_com(
    workbook: xw.Book,
    *,
    include_default_background: bool,
    ignore_colors: set[str] | None,
) -> WorkbookColorsMap | None:
    """Extract colors_map via COM; logs and skips on failure.

    Args:
        workbook: xlwings workbook.
        include_default_background: Whether to include default background.
        ignore_colors: Optional set of color keys to ignore.

    Returns:
        Workbook colors map or None when extraction fails.
    """
    try:
        return extract_sheet_colors_map_com(
            workbook,
            include_default_background=include_default_background,
            ignore_colors=ignore_colors,
        )
    except Exception as exc:
        logger.warning(
            "COM color map extraction failed; falling back to openpyxl. (%r)", exc
        )
        return None


def extract_auto_page_breaks(workbook: xw.Book) -> PrintAreaData:
    """Compute auto page-break rectangles per sheet using Excel COM.

    Args:
        workbook: xlwings workbook.

    Returns:
        Mapping of sheet name to auto page-break areas.
    """
    from typing import Any, cast

    results: PrintAreaData = {}
    for sheet in workbook.sheets:
        try:
            ws_api = cast(Any, sheet.api)
            original_display: bool | None = ws_api.DisplayPageBreaks
            ws_api.DisplayPageBreaks = True
            print_area = ws_api.PageSetup.PrintArea or ws_api.UsedRange.Address
            parts_raw = _split_csv_respecting_quotes(str(print_area))
            area_parts: list[str] = []
            for part in parts_raw:
                rng = _normalize_area_for_sheet(part, sheet.name)
                if rng:
                    area_parts.append(rng)
            hpb = cast(Any, ws_api.HPageBreaks)
            vpb = cast(Any, ws_api.VPageBreaks)
            h_break_rows = [
                hpb.Item(i).Location.Row for i in range(1, int(hpb.Count) + 1)
            ]
            v_break_cols = [
                vpb.Item(i).Location.Column for i in range(1, int(vpb.Count) + 1)
            ]
            for addr in area_parts:
                range_obj = cast(Any, ws_api.Range(addr))
                min_row = int(range_obj.Row)
                max_row = min_row + int(range_obj.Rows.Count) - 1
                min_col = int(range_obj.Column)
                max_col = min_col + int(range_obj.Columns.Count) - 1
                rows = (
                    [min_row]
                    + [r for r in h_break_rows if min_row < r <= max_row]
                    + [max_row + 1]
                )
                cols = (
                    [min_col]
                    + [c for c in v_break_cols if min_col < c <= max_col]
                    + [max_col + 1]
                )
                for i in range(len(rows) - 1):
                    r1, r2 = rows[i], rows[i + 1] - 1
                    for j in range(len(cols) - 1):
                        c1, c2 = cols[j], cols[j + 1] - 1
                        results.setdefault(sheet.name, []).append(
                            PrintArea(r1=r1, c1=c1 - 1, r2=r2, c2=c2 - 1)
                        )
            if original_display is not None:
                ws_api.DisplayPageBreaks = original_display
        except Exception:
            try:
                if original_display is not None:
                    ws_api.DisplayPageBreaks = original_display
            except Exception:
                pass
            continue
    return results


def build_cells_tables_workbook(
    *,
    inputs: ExtractionInputs,
    artifacts: ExtractionArtifacts,
    reason: str,
) -> WorkbookData:
    """Build a WorkbookData containing cells + table_candidates (fallback).

    Args:
        inputs: Pipeline inputs.
        artifacts: Collected artifacts from extraction steps.
        reason: Reason to log for fallback.

    Returns:
        WorkbookData constructed from cells and detected tables.
    """
    colors_map_data = artifacts.colors_map_data
    if inputs.include_colors_map and colors_map_data is None:
        try:
            colors_map_data = extract_sheet_colors_map(
                inputs.file_path,
                include_default_background=inputs.include_default_background,
                ignore_colors=inputs.ignore_colors,
            )
        except Exception as exc:
            logger.warning(
                "Color map extraction failed; skipping colors_map. (%r)", exc
            )
            colors_map_data = None
    sheets: dict[str, SheetData] = {}
    for sheet_name, rows in artifacts.cell_data.items():
        sheet_colors = (
            colors_map_data.get_sheet(sheet_name) if colors_map_data else None
        )
        try:
            tables = detect_tables_openpyxl(inputs.file_path, sheet_name)
        except Exception:
            tables = []
        sheets[sheet_name] = SheetData(
            rows=rows,
            shapes=[],
            charts=[],
            table_candidates=tables,
            print_areas=artifacts.print_area_data.get(sheet_name, [])
            if inputs.include_print_areas
            else [],
            auto_print_areas=[],
            colors_map=sheet_colors.colors_map if sheet_colors else {},
        )
    logger.warning(
        "%s Falling back to cells+tables only; shapes and charts will be empty.",
        reason,
    )
    return WorkbookData(book_name=inputs.file_path.name, sheets=sheets)


def _normalize_area_for_sheet(part: str, ws_name: str) -> str | None:
    """Strip sheet name from a range part when it matches the target sheet.

    Args:
        part: Raw range string part.
        ws_name: Target worksheet name.

    Returns:
        Range without sheet prefix, or None if not matching.
    """
    s = part.strip()
    if "!" not in s:
        return s
    sheet, rng = s.rsplit("!", 1)
    sheet = sheet.strip()
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    return rng if sheet == ws_name else None


def _split_csv_respecting_quotes(raw: str) -> list[str]:
    """Split a CSV-like string while keeping commas inside single quotes intact.

    Args:
        raw: Raw CSV-like string.

    Returns:
        List of split parts.
    """
    parts: list[str] = []
    buf: list[str] = []
    in_quote = False
    i = 0
    while i < len(raw):
        ch = raw[i]
        if ch == "'":
            if in_quote and i + 1 < len(raw) and raw[i + 1] == "'":
                buf.append("''")
                i += 2
                continue
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue
        if ch == "," and not in_quote:
            parts.append("".join(buf).strip())
            buf = []
            i += 1
            continue
        buf.append(ch)
        i += 1
    if buf:
        parts.append("".join(buf).strip())
    return [p for p in parts if p]
